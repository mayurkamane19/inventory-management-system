from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, F, ExpressionWrapper, DecimalField, Count
from django.utils import timezone
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from io import BytesIO

from apps.products.models import Product, Category
from apps.contacts.models import Supplier, Customer
from apps.transactions.models import Purchase, Sale, SaleItem

def filter_by_period(queryset, period, date_field='created_at'):
    now = timezone.now()
    if period == 'daily':
        return queryset.filter(**{f"{date_field}__date": now.date()})
    elif period == 'weekly':
        week_start = now - datetime.timedelta(days=now.weekday())
        return queryset.filter(**{f"{date_field}__gte": week_start})
    elif period == 'monthly':
        return queryset.filter(**{f"{date_field}__year": now.year, f"{date_field}__month": now.month})
    elif period == 'yearly':
        return queryset.filter(**{f"{date_field}__year": now.year})
    return queryset

@login_required
def report_dashboard_view(request):
    report_type = request.GET.get('type', 'sales')
    period = request.GET.get('period', 'monthly')

    sales = filter_by_period(Sale.objects.all(), period, 'sale_date').order_by('-sale_date')
    purchases = filter_by_period(Purchase.objects.all(), period, 'purchase_date').order_by('-purchase_date')
    products = Product.objects.select_related('category').all()

    # Calculations for Profit Report
    total_sales_rev = sales.aggregate(s=Sum('total_amount'))['s'] or 0.00
    total_purchases_cost = purchases.aggregate(s=Sum('total_amount'))['s'] or 0.00
    gross_profit = total_sales_rev - total_purchases_cost

    context = {
        'report_type': report_type,
        'period': period,
        'sales': sales[:20],
        'purchases': purchases[:20],
        'products': products[:20],
        'total_sales_rev': total_sales_rev,
        'total_purchases_cost': total_purchases_cost,
        'gross_profit': gross_profit,
    }
    return render(request, 'reports/report_dashboard.html', context)

@login_required
def export_excel_view(request):
    report_type = request.GET.get('type', 'sales')
    period = request.GET.get('period', 'all')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{report_type.capitalize()} Report"

    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')

    if report_type == 'sales':
        ws.append(['Invoice No', 'Customer', 'Subtotal (₹)', 'Discount (₹)', 'GST (₹)', 'Total Amount (₹)', 'Date'])
        sales = filter_by_period(Sale.objects.all(), period, 'sale_date')
        for s in sales:
            ws.append([s.invoice_no, s.customer.name if s.customer else 'Walk-in', float(s.subtotal), float(s.discount), float(s.gst_amount), float(s.total_amount), s.sale_date.strftime("%Y-%m-%d %H:%M")])
    elif report_type == 'purchases':
        ws.append(['Invoice No', 'Supplier', 'Total Amount (₹)', 'Purchase Date'])
        purchases = filter_by_period(Purchase.objects.all(), period, 'purchase_date')
        for p in purchases:
            ws.append([p.invoice_no, p.supplier.company or p.supplier.name, float(p.total_amount), p.purchase_date.strftime("%Y-%m-%d %H:%M")])
    else: # products
        ws.append(['SKU', 'Product Name', 'Category', 'Selling Price (₹)', 'Stock Quantity', 'Stock Value (₹)'])
        for pr in Product.objects.all():
            ws.append([pr.sku, pr.name, pr.category.name, float(pr.selling_price), pr.stock_quantity, float(pr.inventory_value)])

    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report.xlsx"'
    wb.save(response)
    return response

@login_required
def export_pdf_view(request):
    report_type = request.GET.get('type', 'sales')
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(name='TitleStyle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1F4E79'))
    elements.append(Paragraph(f"Inventory System - {report_type.capitalize()} Report", title_style))
    elements.append(Spacer(1, 12))

    data = []
    if report_type == 'sales':
        data.append(['Invoice', 'Customer', 'Total (₹)', 'Date'])
        sales = Sale.objects.all().order_by('-sale_date')[:30]
        for s in sales:
            data.append([s.invoice_no, s.customer.name if s.customer else 'Walk-in', f"₹{s.total_amount}", s.sale_date.strftime("%Y-%m-%d")])
    elif report_type == 'purchases':
        data.append(['Invoice', 'Supplier', 'Total (₹)', 'Date'])
        purchases = Purchase.objects.all().order_by('-purchase_date')[:30]
        for p in purchases:
            data.append([p.invoice_no, p.supplier.name, f"₹{p.total_amount}", p.purchase_date.strftime("%Y-%m-%d")])
    else:
        data.append(['SKU', 'Product Name', 'Stock', 'Price (₹)'])
        for pr in Product.objects.all()[:30]:
            data.append([pr.sku, pr.name, str(pr.stock_quantity), f"₹{pr.selling_price}"])

    t = Table(data, colWidths=[120, 180, 100, 100])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F2F2F2')),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#DDDDDD')),
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report.pdf"'
    return response
